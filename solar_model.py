import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 1. Criação do Workbook e Abas
wb = openpyxl.Workbook()
ws_dashboard = wb.active
ws_dashboard.title = "Solar Calculator"
ws_calcs = wb.create_sheet(title = "Daily Calculations")

# Forçar exibição de linhas de grade
ws_dashboard.views.sheetView[0].showGridLines = True
ws_calcs.views.sheetView[0].showGridLines = True

# Paleta de Cores Profissional (Classic Navy)
NAVY_DARK = "1F497D"
NAVY_LIGHT = "DCE6F1"
ZEBRA_FILL = "F2F5F8"
WHITE = "FFFFFF"
GRAY_TEXT = "595959"
BORDER_GRAY = "D9D9D9"

# Estilos de Texto e Preenchimento
font_title = Font(name="Arial", size=16, bold=True, color=WHITE)
font_section = Font(name="Arial", size=12, bold=True, color=NAVY_DARK)
font_header = Font(name="Arial", size=11, bold=True, color=WHITE)
font_data_bold = Font(name="Arial", size=11, bold=True)
font_data = Font(name="Arial", size=11)
font_italic = Font(name="Arial", size=9, italic=True, color=GRAY_TEXT)

fill_title = PatternFill(start_color=NAVY_DARK, end_color=NAVY_DARK, fill_type="solid")
fill_header = PatternFill(start_color=NAVY_DARK, end_color=NAVY_DARK, fill_type="solid")
fill_light = PatternFill(start_color=NAVY_LIGHT, end_color=NAVY_LIGHT, fill_type="solid")
fill_zebra = PatternFill(start_color=ZEBRA_FILL, end_color=ZEBRA_FILL, fill_type="solid")

thin_border_side = Side(border_style="thin", color=BORDER_GRAY)
thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")

# --- ABA 1: PAINEL DE DIAGNÓSTICO (DASHBOARD) ---
ws_dashboard.merge_cells("A1:G2")
title_cell = ws_dashboard["A1"]
title_cell.value = "SOLAR MODEL - ERBS & HAY TRANSPOSITION ENGINE"
title_cell.font = font_title
title_cell.fill = fill_title
title_cell.alignment = align_center

ws_dashboard["A4"] = "INPUT PARAMETERS"
ws_dashboard["A4"].font = font_section

inputs = [
    ("Latitude [°]", -23.55, "B5", "Negativo para Sul, Positivo para Norte (ex: SP = -23.55)"),
    ("Longitude [°]", -46.63, "B6", "Negativo para Oeste, Positivo para Leste (ex: SP = -46.63)"),
    ("Time Zone [h]", -3, "B7", "Fuso horário relativo a GMT/UTC (ex: Brasília = -3)"),
    ("Day of the Year [NoDay]", 172, "B8", "Dia sequencial de 1 a 365 (ex: 172 = Solstício de Inverno)"),
    ("Year", 2026, "B9", "Ano de análise (identifica bissexto automaticamente)"),
    ("Legal Time [h]", 12.0, "B10", "Hora legal do dia em formato decimal (ex: 13.5 = 13h30)"),
    ("Plane Tilt (TiltPl) [°]", 23.5, "B11", "Ângulo de inclinação do painel em relação ao solo"),
    ("Plane Azimuth (AzimPl) [°]", 0.0, "B12", "Orientação: 0=Norte (Hem. Sul) ou Sul (Hem. Norte)"),
    ("Peak GHI [W/m²]", 800, "B13", "Irradiância Global Horizontal de referência ao meio-dia"),
    ("Solar Constant Esc [W/m²]", 1361, "B14", "Constante Solar Extraterrestre Padrão")
]

for label, val, cell, desc in inputs:
    row = int(cell[1:])
    ws_dashboard.cell(row=row, column=1, value=label).font = font_data_bold
    val_cell = ws_dashboard.cell(row=row, column=2, value=val)
    val_cell.font = font_data
    val_cell.alignment = align_right
    val_cell.border = thin_border
    ws_dashboard.cell(row=row, column=3, value=desc).font = font_italic

# Ajustes de colunas do painel
ws_dashboard.column_dimensions['A'].width = 32
ws_dashboard.column_dimensions['B'].width = 15
ws_dashboard.column_dimensions['C'].width = 45
ws_dashboard.column_dimensions['E'].width = 30
ws_dashboard.column_dimensions['F'].width = 15
ws_dashboard.column_dimensions['G'].width = 18

# --- ABA 2: MATRIZ DE CÁLCULO DIÁRIO (24 HORAS) ---
ws_calcs["A1"] = "Perfil Solar Horário Completo - Modelo de Decomposição (Erbs) e Transposição (Hay)"
ws_calcs["A1"].font = Font(name="Arial", size=14, bold=True, color=NAVY_DARK)

headers_calcs = [
    "Legal Time", "Solar Alt [°]", "ENI [W/m²]", "EHI [W/m²]",
    "Dynamic GHI", "Kt (Clearness)", "Kd (Erbs)", "DHI [W/m²]",
    "BHI [W/m²]", "DNI [W/m²]", "Incidence Angle [°]", "Rb (Geo Factor)",
    "A (Anisotropy)", "I_Beam [W/m²]", "I_Diffuse [W/m²]", "I_Albedo [W/m²]", "G_Tilt (TOTAL)"
]

for col_idx, h in enumerate(headers_calcs, start=1):
    c = ws_calcs.cell(row=3, column=col_idx, value=h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = align_center

# Variáveis auxiliares de órbita em células ocultas na aba de cálculo para simplificar fórmulas de 24h
ws_calcs["Z1"] = "=IF(MOD('Solar Calculator'!$B$9,4)=0,366,365)" # NDaysY
ws_calcs["Z2"] = "=79+(MOD('Solar Calculator'!$B$9,4)/4)"        # NoDayOrig
ws_calcs["Z3"] = "=2*PI()*('Solar Calculator'!$B$8-1)/365.25"    # YAngle
ws_calcs["Z4"] = "=0.0072*COS(Z3)-0.0528*COS(2*Z3)-0.1229*SIN(Z3)-0.1565*SIN(2*Z3)" # TE
ws_calcs["Z5"] = "=DEGREES(ASIN(SIN(RADIANS(23.433))*SIN(2*PI()*('Solar Calculator'!$B$8-$Z$2)/$Z$1)))" # Decl

# Loop para gerar as 24 horas do dia
for h_idx in range(0, 24):
    row = 4 + h_idx
    is_zebra = (h_idx % 2 == 1)

    # Hora Legal
    ws_calcs.cell(row=row, column=1, value=float(h_idx)).number_format = "0.0"

    # Ângulo Horário (graus) para esta hora — embutido inline na fórmula da Col 2
    ha_expr = f"15*((A{row}-('Solar Calculator'!$B$7-('Solar Calculator'!$B$6/15)-$Z$4))-12)"

    # Col 2: Altura Solar (HSun) — declina via $Z$5, HA via ha_expr inline
    ws_calcs.cell(row=row, column=2, value=f"=DEGREES(ASIN(-SIN(RADIANS('Solar Calculator'!$B$5))*SIN(RADIANS($Z$5)) + COS(RADIANS('Solar Calculator'!$B$5))*COS(RADIANS($Z$5))*COS(RADIANS({ha_expr}))))").number_format = "0.00"

    # Col 3: ENI
    ws_calcs.cell(row=row, column=3, value="='Solar Calculator'!$B$14*(1+0.033*COS($Z$3))").number_format = "#,##0"

    # Col 4: EHI (Com trava física de horizonte > 0)
    ws_calcs.cell(row=row, column=4, value=f"=IF(B{row}>0, C{row}*SIN(RADIANS(B{row})), 0)").number_format = "#,##0"

    # Col 5: Dynamic GHI de superfície
    ws_calcs.cell(row=row, column=5, value=f"=IF(B{row}>0, 'Solar Calculator'!$B$13*SIN(RADIANS(B{row})), 0)").number_format = "#,##0"

    # Col 6: Índice de Clareza Kt (Com trava de segurança < 2 graus)
    ws_calcs.cell(row=row, column=6, value=f"=IF(B{row}<2, 0, IF(D{row}=0, 0, E{row}/D{row}))").number_format = "0.000"

    # Col 7: Fração Difusa Kd (Modelo de Erbs 1982)
    ws_calcs.cell(row=row, column=7, value=f"=IF(F{row}<=0, 0, IF(F{row}<=0.22, 1-0.09*F{row}, IF(F{row}<=0.8, 0.9511-0.1604*F{row}+4.388*F{row}^2-16.638*F{row}^3+12.336*F{row}^4, 0.165)))").number_format = "0.000"

    # Col 8: DHI (Difusa Horizontal)
    ws_calcs.cell(row=row, column=8, value=f"=G{row}*E{row}").number_format = "#,##0"

    # Col 9: BHI (Direta Horizontal)
    ws_calcs.cell(row=row, column=9, value=f"=E{row}-H{row}").number_format = "#,##0"

    # Col 10: DNI (Direta Normal)
    ws_calcs.cell(row=row, column=10, value=f"=IF(B{row}>0, I{row}/SIN(RADIANS(B{row})), 0)").number_format = "#,##0"

    # Col 11: Ângulo de Incidência do Painel (IA)
    ws_calcs.cell(row=row, column=11, value=f"=IF(B{row}>0, DEGREES(ACOS(COS(RADIANS(0-'Solar Calculator'!$B$12))*COS(RADIANS(B{row}))*SIN(RADIANS('Solar Calculator'!$B$11)) + SIN(RADIANS(B{row}))*COS(RADIANS('Solar Calculator'!$B$11)))), 90)").number_format = "0.00"

    # === MODELO DE TRANSPOSIÇÃO DE HAY (1980) ===
    # Col 12: Fator Geométrico Rb
    ws_calcs.cell(row=row, column=12, value=f"=IF(B{row}>0, COS(RADIANS(K{row}))/SIN(RADIANS(B{row})), 0)").number_format = "0.0000"

    # Col 13: Índice de Anisotropia A
    ws_calcs.cell(row=row, column=13, value=f"=IF(D{row}=0, 0, I{row}/D{row})").number_format = "0.0000"

    # Col 14: Componente Direta Inclinada (I_Beam)
    ws_calcs.cell(row=row, column=14, value=f"=I{row}*L{row}").number_format = "#,##0"

    # Col 15: Componente Difusa Inclinada Anisotrópica de Hay (I_Diffuse)
    ws_calcs.cell(row=row, column=15, value=f"=H{row}*(M{row}*L{row} + (1-M{row})*(1+COS(RADIANS('Solar Calculator'!$B$11)))/2)").number_format = "#,##0"

    # Col 16: Componente de Albedo/Solo (I_Albedo com rho=0.2)
    ws_calcs.cell(row=row, column=16, value=f"=E{row}*0.2*(1-COS(RADIANS('Solar Calculator'!$B$11)))/2").number_format = "#,##0"

    # Col 17: IRRADIÂNCIA GLOBAL TOTAL NO PAINEL INCLINADO (G_Tilt)
    ws_calcs.cell(row=row, column=17, value=f"=N{row}+O{row}+P{row}").number_format = "#,##0"

    # Formatação visual das células da linha
    for col_c in range(1, 18):
        cell_obj = ws_calcs.cell(row=row, column=col_c)
        cell_obj.font = font_data
        cell_obj.border = thin_border
        cell_obj.alignment = align_right
        if is_zebra:
            cell_obj.fill = fill_zebra
        if col_c == 17:
            cell_obj.fill = fill_light
            cell_obj.font = font_data_bold

# Auto-ajuste de largura de colunas
for col in ws_calcs.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws_calcs.column_dimensions[col_letter].width = max(max_len + 2, 14)

# Salvar Arquivo final
file_name = "Solar_Model_Hay_Transposition.xlsx"
wb.save(file_name)
print(f"Sucesso! Planilha criada localmente: {file_name}")
