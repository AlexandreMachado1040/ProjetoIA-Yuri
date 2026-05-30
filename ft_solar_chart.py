"""
============================================================
 Fator de Transposição Solar — Modelo Hay + Erbs  v2.0
============================================================
 Modos de uso:
   python ft_solar_chart.py                              # prompt interativo
   python ft_solar_chart.py --xlsx arquivo.xlsx          # planilha existente
   python ft_solar_chart.py --cidade "São Paulo"         # geocoding + Open-Meteo
   python ft_solar_chart.py --lat -23.55 --lon -46.63   # coordenadas diretas

 Flags opcionais (combinam com qualquer modo):
   --dia    172    Dia do ano (1-365), padrão = hoje
   --hora   12.0   Hora legal decimal
   --tilt   23.5   Inclinação atual do painel [°]
   --azim   0.0    Azimute atual do painel [°]  (0=Sul no HS)
   --ghi    800    GHI manual [W/m²]  (ignora API)
   --albedo 0.2    Albedo da superfície
   --api    auto|openmeteo|nasapower|none
============================================================
"""

import sys
import os
import math
import argparse
import datetime
import json
import urllib.request as _urlreq
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap
from scipy.interpolate import RegularGridInterpolator
from scipy.ndimage import gaussian_filter

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

PI = math.pi


# ─────────────────────────────────────────────────────────────────────────────
#  GEOCODIFICAÇÃO — Nominatim / OpenStreetMap (sem chave de API)
# ─────────────────────────────────────────────────────────────────────────────
def geocodificar(nome: str):
    """Converte nome de cidade em (lat, lon, display_name) ou None."""
    import urllib.parse
    q = urllib.parse.quote(nome)
    url = f"https://nominatim.openstreetmap.org/search?q={q}&format=json&limit=1"
    try:
        req = _urlreq.Request(url, headers={"User-Agent": "ft_solar_chart/2.0"})
        with _urlreq.urlopen(req, timeout=8) as r:
            dados = json.loads(r.read())
        if dados:
            return float(dados[0]["lat"]), float(dados[0]["lon"]), dados[0]["display_name"]
    except Exception as e:
        print(f"  Geocoding falhou: {e}")
    return None


# ─────────────────────────────────────────────────────────────────────────────
#  POSIÇÃO SOLAR — cálculo interno, sem Excel
# ─────────────────────────────────────────────────────────────────────────────
def posicao_solar(lat, lon, tzone, no_day, hora_leg, ano=None):
    """
    Retorna dict com HSun [°], AzSun [°] e ENI [W/m²].
    Convenção AzSun: 0=Sul, +Leste, -Oeste (hemisf. sul padrão fotovoltaico).
    Usa Equação do Tempo de Spencer (1971) e Declinação de Cooper (1969).
    """
    if ano is None:
        ano = datetime.date.today().year

    B = 2 * PI * (no_day - 1) / 365.25

    # Equação do tempo [min]
    ET_min = 229.18 * (
        0.000075
        + 0.001868 * math.cos(B)   - 0.032077 * math.sin(B)
        - 0.014615 * math.cos(2*B) - 0.04089  * math.sin(2*B)
    )

    # Hora solar verdadeira [h]
    hora_solar = hora_leg + (lon / 15.0 - tzone) + ET_min / 60.0

    # Ângulo horário [°] — negativo de manhã, positivo à tarde
    HA_deg = 15.0 * (hora_solar - 12.0)
    HA_r   = math.radians(HA_deg)

    # Declinação solar [°] — Cooper 1969
    decl_r = math.radians(23.433 * math.sin(2 * PI * (no_day - 80) / 365.0))
    lat_r  = math.radians(lat)

    # Altitude solar [°]
    sin_h = (math.sin(lat_r) * math.sin(decl_r)
             + math.cos(lat_r) * math.cos(decl_r) * math.cos(HA_r))
    HSun = math.degrees(math.asin(max(-1.0, min(1.0, sin_h))))

    # Azimute solar [°]
    AzSun = 0.0
    if HSun > 0.5:
        cos_az = ((math.sin(decl_r) * math.cos(lat_r)
                   - math.cos(decl_r) * math.sin(lat_r) * math.cos(HA_r))
                  / math.cos(math.radians(HSun)))
        AzSun = math.degrees(math.acos(max(-1.0, min(1.0, cos_az))))
        if HA_deg > 0:   # tarde → oeste → negativo
            AzSun = -AzSun

    # ENI [W/m²]
    ENI = 1361.0 * (1 + 0.033 * math.cos(2 * PI * (no_day - 1) / 365.0))

    return {"HSun": HSun, "AzSun": AzSun, "ENI": ENI}


# ─────────────────────────────────────────────────────────────────────────────
#  DECOMPOSIÇÃO ERBS (1982) — GHI → DNI + DHI
# ─────────────────────────────────────────────────────────────────────────────
def decompor_erbs(GHI, HSun, ENI):
    """Retorna (DNI, DHI). Requer GHI > 0 e HSun > 2°."""
    if HSun <= 2 or GHI <= 0:
        return 0.0, 0.0
    EHI = ENI * math.sin(math.radians(HSun))
    if EHI <= 0:
        return 0.0, 0.0
    Kt = min(1.0, max(0.0, GHI / EHI))
    if Kt <= 0.22:
        Kd = 1 - 0.09 * Kt
    elif Kt <= 0.80:
        Kd = (0.9511 - 0.1604*Kt + 4.388*Kt**2
              - 16.638*Kt**3 + 12.336*Kt**4)
    else:
        Kd = 0.165
    DHI = Kd * GHI
    DNI = (GHI - DHI) / math.sin(math.radians(HSun))
    return max(0.0, DNI), max(0.0, DHI)


# ─────────────────────────────────────────────────────────────────────────────
#  APIS DE IRRADIÂNCIA
# ─────────────────────────────────────────────────────────────────────────────
def _data_para_api(no_day, ano=None):
    """Retorna date para o dia do ano. Se for data futura, recua um ano."""
    hoje = datetime.date.today()
    ano  = ano or hoje.year
    d = datetime.date(ano, 1, 1) + datetime.timedelta(days=no_day - 1)
    if d > hoje:
        d = datetime.date(ano - 1, 1, 1) + datetime.timedelta(days=no_day - 1)
    return d


def buscar_openmeteo(lat, lon, no_day, hora_leg):
    """
    Busca GHI, DNI, DHI via Open-Meteo Archive API.
    Gratuito, sem chave. Retorna (GHI, DNI, DHI) ou None.
    """
    d  = _data_para_api(no_day)
    ds = d.strftime("%Y-%m-%d")
    h  = int(hora_leg)
    url = (
        "https://archive-api.open-meteo.com/v1/archive"
        f"?latitude={lat:.4f}&longitude={lon:.4f}"
        f"&start_date={ds}&end_date={ds}"
        "&hourly=shortwave_radiation,diffuse_radiation,direct_normal_irradiance"
        "&timezone=auto"
    )
    try:
        req = _urlreq.Request(url, headers={"User-Agent": "ft_solar_chart/2.0"})
        with _urlreq.urlopen(req, timeout=12) as r:
            data = json.loads(r.read())
        hly   = data["hourly"]
        times = hly["time"]
        target = f"{ds}T{h:02d}:00"
        idx    = times.index(target) if target in times else h
        GHI = max(0.0, float(hly["shortwave_radiation"][idx] or 0))
        DHI = max(0.0, float(hly["diffuse_radiation"][idx] or 0))
        DNI = max(0.0, float(hly["direct_normal_irradiance"][idx] or 0))
        print(f"  Open-Meteo [{ds} {h:02d}h]: GHI={GHI:.0f}  DNI={DNI:.0f}  DHI={DHI:.0f} W/m2")
        return GHI, DNI, DHI
    except Exception as e:
        print(f"  Open-Meteo falhou: {e}")
        return None


def buscar_nasapower(lat, lon, no_day, hora_leg):
    """
    Busca GHI, DNI, DHI via NASA POWER Hourly API.
    Gratuito, sem chave. Retorna (GHI, DNI, DHI) ou None.
    """
    d  = _data_para_api(no_day)
    ds = d.strftime("%Y%m%d")
    h  = int(hora_leg)
    url = (
        "https://power.larc.nasa.gov/api/temporal/hourly/point"
        "?parameters=ALLSKY_SFC_SW_DWN,ALLSKY_SFC_SW_DIFF,ALLSKY_SFC_SW_DNI"
        f"&community=RE&longitude={lon:.4f}&latitude={lat:.4f}"
        f"&start={ds}&end={ds}&format=JSON"
    )
    try:
        req = _urlreq.Request(url, headers={"User-Agent": "ft_solar_chart/2.0"})
        with _urlreq.urlopen(req, timeout=25) as r:
            data = json.loads(r.read())
        par = data["properties"]["parameter"]

        def _get(key):
            d_par = par.get(key, {})
            # Chave horária NASA POWER: "YYYYMMDDHHNN" (HH=hora, NN=30)
            for chave in [f"{ds}{h:02d}30", f"{ds}{h:02d}00"]:
                if chave in d_par:
                    v = float(d_par[chave])
                    return max(0.0, v) if v != -999.0 else 0.0
            # Fallback: varre chaves pela hora
            for chave, v in d_par.items():
                if len(chave) >= 10 and chave[8:10] == f"{h:02d}":
                    v = float(v)
                    return max(0.0, v) if v != -999.0 else 0.0
            return 0.0

        GHI = _get("ALLSKY_SFC_SW_DWN")
        DHI = _get("ALLSKY_SFC_SW_DIFF")
        DNI = _get("ALLSKY_SFC_SW_DNI")
        print(f"  NASA POWER [{d} {h:02d}h]: GHI={GHI:.0f}  DNI={DNI:.0f}  DHI={DHI:.0f} W/m2")
        return GHI, DNI, DHI
    except Exception as e:
        print(f"  NASA POWER falhou: {e}")
        return None


def buscar_irradiancia(lat, lon, no_day, hora_leg, modo="auto"):
    """Orquestra as chamadas de API. modo = auto|openmeteo|nasapower|none."""
    if modo == "none":
        return None
    if modo in ("auto", "openmeteo"):
        res = buscar_openmeteo(lat, lon, no_day, hora_leg)
        if res:
            return res
    if modo in ("auto", "nasapower"):
        res = buscar_nasapower(lat, lon, no_day, hora_leg)
        if res:
            return res
    return None


# ─────────────────────────────────────────────────────────────────────────────
#  LER PLANILHA EXISTENTE
# ─────────────────────────────────────────────────────────────────────────────
def ler_planilha(caminho: str) -> dict:
    """Lê parâmetros da aba 'Solar Calculator' gerada pelo solar_model.py."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl não instalado.")
    wb = load_workbook(caminho, data_only=True)
    aba_candidatas = ["Solar Calculator", "Dashboard", wb.sheetnames[0]]
    ws = next((wb[n] for n in aba_candidatas if n in wb.sheetnames), wb.active)

    def cel(addr, default=0.0):
        v = ws[addr].value
        try:
            return float(v)
        except (TypeError, ValueError):
            return default

    params = {
        "lat":       cel("B5"),
        "lon":       cel("B6"),
        "tzone":     int(cel("B7")),
        "no_day":    int(cel("B8")),
        "hora_leg":  cel("B10"),
        "tilt":      cel("B11"),
        "azimute":   cel("B12"),
        "albedo":    0.2,
        "nome_local": os.path.splitext(os.path.basename(caminho))[0],
        # Resultados calculados — podem não existir na planilha nova
        "HSun":  cel("G10", None),
        "AzSun": cel("G11", None),
        "GHI":   cel("G16", None),
        "DNI":   cel("G21", None),
        "DHI":   cel("G19", None),
        "ENI":   cel("G14", None),
    }
    wb.close()
    return params


# ─────────────────────────────────────────────────────────────────────────────
#  MONTAGEM DO DICT COMPLETO DE PARÂMETROS
# ─────────────────────────────────────────────────────────────────────────────
def montar_params(base: dict, modo_api="auto") -> dict:
    """
    Completa campos faltantes: calcula posição solar internamente e
    busca irradiância via API quando GHI não foi fornecido.
    """
    p = dict(base)

    # Posição solar — sempre recalculada para garantir consistência
    sol = posicao_solar(p["lat"], p["lon"], p["tzone"], p["no_day"], p["hora_leg"])
    for k in ("HSun", "AzSun", "ENI"):
        if not p.get(k):
            p[k] = sol[k]

    # Irradiância — API se GHI ausente ou zero
    if not p.get("GHI") or p["GHI"] <= 0:
        irr = buscar_irradiancia(p["lat"], p["lon"], p["no_day"], p["hora_leg"], modo_api)
        if irr:
            p["GHI"], p["DNI"], p["DHI"] = irr
        else:
            # Fallback: estima GHI como 75% do EHI (céu médio)
            EHI = p["ENI"] * math.sin(math.radians(max(0.0, p["HSun"])))
            p["GHI"] = max(0.0, 0.75 * EHI)
            print(f"  GHI estimado (75% EHI): {p['GHI']:.0f} W/m2")

    # Decomposição Erbs se DNI/DHI ausentes
    if not p.get("DNI") or not p.get("DHI"):
        p["DNI"], p["DHI"] = decompor_erbs(p["GHI"], p["HSun"], p["ENI"])

    p.setdefault("tilt",       23.5)
    p.setdefault("azimute",    0.0)
    p.setdefault("albedo",     0.2)
    p.setdefault("nome_local", f"Lat={p['lat']:.2f} Lon={p['lon']:.2f}")
    return p


# ─────────────────────────────────────────────────────────────────────────────
#  CÁLCULO FT — Modelo Hay (1980)
# ─────────────────────────────────────────────────────────────────────────────
def calcular_ft(HSun, AzSun, GHI, DNI, DHI, ENI, albedo, tilt_deg, az_deg):
    if HSun <= 0 or GHI <= 0:
        return 0.0
    h_r    = math.radians(HSun)
    t_r    = math.radians(tilt_deg)
    az_rel = math.radians(AzSun - az_deg)
    cos_ia = (math.cos(az_rel) * math.cos(h_r) * math.sin(t_r)
              + math.sin(h_r) * math.cos(t_r))
    Rb     = max(0.0, cos_ia / math.sin(h_r))   # clamped: nunca negativo
    fAni   = max(0.0, min(1.0, DNI / ENI)) if ENI > 0 else 0.0
    Ibeam  = DNI * Rb
    Idiff  = DHI * (fAni * Rb + (1 - fAni) * (1 + math.cos(t_r)) / 2)
    Ialb   = GHI * albedo * (1 - math.cos(t_r)) / 2
    return (Ibeam + Idiff + Ialb) / GHI


def calcular_matriz_ft(p: dict):
    """Grade 5°×5°: 37 azimutes × 19 inclinações = 703 combinações."""
    azimutes    = np.arange(-90, 91, 5, dtype=float)
    inclinacoes = np.arange(0,  91, 5, dtype=float)
    ft_matrix   = np.zeros((len(inclinacoes), len(azimutes)))
    for j, tilt in enumerate(inclinacoes):
        for i, az in enumerate(azimutes):
            ft_matrix[j, i] = calcular_ft(
                p["HSun"], p["AzSun"], p["GHI"], p["DNI"],
                p["DHI"],  p["ENI"],  p["albedo"], tilt, az
            )
    return azimutes, inclinacoes, ft_matrix


# ─────────────────────────────────────────────────────────────────────────────
#  GRÁFICO
# ─────────────────────────────────────────────────────────────────────────────
def gerar_grafico(p, azimutes, inclinacoes, ft_matrix, caminho_saida):
    az_fine  = np.linspace(-90, 90, 721)
    tlt_fine = np.linspace(0,   90, 361)
    AZ_g, TL_g = np.meshgrid(az_fine, tlt_fine)

    interp = RegularGridInterpolator(
        (inclinacoes, azimutes), ft_matrix,
        method="linear", bounds_error=False, fill_value=None
    )
    FT_fine = interp(np.column_stack([TL_g.ravel(), AZ_g.ravel()])).reshape(TL_g.shape)
    FT_fine = gaussian_filter(FT_fine, sigma=1.2)

    best_val  = float(ft_matrix.max())
    best_idx  = np.unravel_index(ft_matrix.argmax(), ft_matrix.shape)
    best_tilt = float(inclinacoes[best_idx[0]])
    best_az   = float(azimutes[best_idx[1]])
    ft_atual  = float(interp(np.array([[p["tilt"], p["azimute"]]]))[0])

    cmap = LinearSegmentedColormap.from_list("pvstyle", [
        (0.10, 0.20, 0.60), (0.30, 0.55, 0.85), (0.70, 0.85, 0.97),
        (1.00, 1.00, 1.00), (1.00, 0.88, 0.60), (1.00, 0.55, 0.15),
        (0.85, 0.10, 0.10),
    ], N=512)

    fig, ax = plt.subplots(figsize=(12, 9), dpi=150)
    fig.patch.set_facecolor("white")
    ax.set_facecolor("#EFF4F9")

    vmin, vmax = 0.20, 1.10
    cf = ax.contourf(AZ_g, TL_g, FT_fine,
                     levels=np.linspace(vmin, vmax, 200),
                     cmap=cmap, vmin=vmin, vmax=vmax, extend="both")

    niveis_iso = [0.5, 0.6, 0.7, 0.8, 0.9, 1.0, 1.02, 1.04]
    ct = ax.contour(AZ_g, TL_g, FT_fine, levels=niveis_iso,
                    colors="black", linewidths=0.7, linestyles="--", alpha=0.75)
    ax.clabel(ct, fmt={lv: f"FT={lv:.2f}" for lv in niveis_iso},
              fontsize=8, inline=True, inline_spacing=6)

    # Linha vermelha: orientação ótima por inclinação
    opt_az = []
    for tilt in inclinacoes:
        row_idx    = int(np.argmin(np.abs(tlt_fine - tilt)))
        az_opt_idx = int(np.argmax(FT_fine[row_idx, :]))
        opt_az.append(float(az_fine[az_opt_idx]))
    ax.plot(opt_az, inclinacoes, "-", color="crimson", lw=3.2, zorder=9,
            label="Orientação ótima por inclinação")

    # Ponto ótimo global
    ax.plot(best_az, best_tilt, "o", markersize=13, color="darkred",
            markeredgecolor="white", markeredgewidth=2, zorder=12)
    ax.annotate(
        f"FT max = {best_val:.3f}\nTilt={best_tilt:.0f} / Az={best_az:+.0f}",
        xy=(best_az, best_tilt), xytext=(best_az + 10, best_tilt - 12),
        fontsize=9, color="darkred", fontweight="bold",
        bbox=dict(boxstyle="round,pad=0.3", fc="white", alpha=0.88, ec="darkred"),
        arrowprops=dict(arrowstyle="->", color="darkred", lw=1.5)
    )

    # Ponto seleção atual
    ax.plot(p["azimute"], p["tilt"], "s", markersize=11, color="crimson",
            markeredgecolor="white", markeredgewidth=2, zorder=13)
    ax.annotate(
        f"Selecao atual\nTilt={p['tilt']:.1f} / Az={p['azimute']:+.0f}\nFT={ft_atual:.3f}",
        xy=(p["azimute"], p["tilt"]), xytext=(p["azimute"] + 14, p["tilt"] + 8),
        fontsize=8.5, color="crimson", fontweight="bold",
        bbox=dict(boxstyle="round,pad=0.3", fc="white", alpha=0.88, ec="crimson"),
        arrowprops=dict(arrowstyle="->", color="crimson", lw=1.4)
    )

    cb = fig.colorbar(cf, ax=ax, pad=0.02, aspect=30, shrink=0.90)
    cb.set_label("Fator de Transposicao  FT = G_tilt / GHI", fontsize=10)
    cb.ax.tick_params(labelsize=8.5)
    cb.set_ticks([0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0, 1.05])

    ax.set_xlim(-90, 90)
    ax.set_ylim(0, 90)
    ax.set_xticks(range(-90, 91, 15))
    ax.set_yticks(range(0, 91, 5))
    ax.tick_params(labelsize=9)
    ax.grid(True, color="white", linewidth=0.6, alpha=0.7)
    ax.set_xlabel("<- Leste  |  Azimute do plano [0=Sul]  |  Oeste ->", fontsize=10)
    ax.set_ylabel("Inclinacao do plano [graus]", fontsize=10)
    ax.set_title(
        f"Mapa do Fator de Transposicao  -  Modelo Hay + Erbs\n"
        f"{p['nome_local']}  |  Dia={p['no_day']}"
        f"  |  Hora={p['hora_leg']}h  |  Alt. Solar={p['HSun']:.1f}",
        fontsize=12, fontweight="bold", pad=12
    )
    ax.legend(loc="upper left", fontsize=9, framealpha=0.92,
              edgecolor="crimson", facecolor="white")
    ax.text(0.995, 0.01,
            "Erbs (1982) + Hay (1980)  |  FT = (Ibeam + Idiff_Hay + Ialbedo) / GHI",
            transform=ax.transAxes, fontsize=7, color="#666", ha="right", va="bottom")

    plt.tight_layout()
    plt.savefig(caminho_saida, dpi=150, bbox_inches="tight",
                facecolor="white", edgecolor="none")
    plt.close()
    return ft_atual, best_val, best_tilt, best_az


# ─────────────────────────────────────────────────────────────────────────────
#  RELATORIO TERMINAL
# ─────────────────────────────────────────────────────────────────────────────
def imprimir_relatorio(p, ft_atual, best_val, best_tilt, best_az, caminho_img):
    sep = "-" * 60
    print(f"\n{sep}")
    print("  FATOR DE TRANSPOSICAO - Modelo Hay + Erbs")
    print(sep)
    print(f"  Local             : {p['nome_local']}")
    print(f"  Latitude          : {p['lat']}")
    print(f"  Dia do ano        : {p['no_day']}")
    print(f"  Hora legal        : {p['hora_leg']} h")
    print(f"  Altitude solar    : {p['HSun']:.2f} graus")
    print(f"  Azimute solar     : {p['AzSun']:.2f} graus")
    print(sep)
    print(f"  GHI               : {p['GHI']:.1f} W/m2")
    print(f"  DNI               : {p['DNI']:.1f} W/m2")
    print(f"  DHI               : {p['DHI']:.1f} W/m2")
    print(f"  Albedo            : {p['albedo']:.2f}")
    print(sep)
    print(f"  Selecao atual     : Tilt={p['tilt']:.1f}  Az={p['azimute']:+.0f}")
    print(f"  FT atual          : {ft_atual:.4f}")
    print(sep)
    print(f"  FT MAXIMO         : {best_val:.4f}")
    print(f"  Orientacao otima  : Tilt={best_tilt:.0f}  Az={best_az:+.0f}")
    print(sep)
    print(f"\n  Grafico salvo em  : {caminho_img}\n")


# ─────────────────────────────────────────────────────────────────────────────
#  INPUT INTERATIVO
# ─────────────────────────────────────────────────────────────────────────────
def _ask(prompt, default, tipo=float):
    try:
        txt = input(f"  {prompt} [{default}]: ").strip()
        return tipo(txt) if txt else tipo(default)
    except (ValueError, EOFError):
        return tipo(default)


def input_interativo() -> dict:
    hoje = datetime.date.today()
    no_day_hoje = hoje.timetuple().tm_yday

    print("\n" + "=" * 60)
    print("  SOLAR FT v2.0 - Entrada de Parametros")
    print("=" * 60)
    print("\n  [1] Buscar cidade por nome (geocoding automatico)")
    print("  [2] Inserir latitude/longitude manualmente")
    modo = input("\n  Escolha [1/2]: ").strip()

    lat = lon = tzone = nome_local = None

    if modo == "1":
        cidade = input("  Nome da cidade: ").strip()
        print(f"  Buscando '{cidade}'...")
        res = geocodificar(cidade)
        if res:
            lat, lon, display = res
            nome_local = display.split(",")[0]
            print(f"  Encontrado: {nome_local}  (lat={lat:.4f}, lon={lon:.4f})")
            tzone_est = round(lon / 15)
            tzone = _ask(f"Fuso horario GMT (estimado={tzone_est})", tzone_est, int)
        else:
            print("  Geocoding falhou. Insira as coordenadas manualmente.")
            modo = "2"

    if modo != "1" or lat is None:
        lat       = _ask("Latitude [negativo=Sul]", -23.55)
        lon       = _ask("Longitude [negativo=Oeste]", -46.63)
        tzone_est = round(lon / 15)
        tzone     = _ask(f"Fuso horario GMT (estimado={tzone_est})", tzone_est, int)
        nome_local = f"Lat={lat:.2f} Lon={lon:.2f}"

    print()
    no_day   = _ask(f"Dia do ano [1-365, hoje={no_day_hoje}]", no_day_hoje, int)
    hora_leg = _ask("Hora legal [decimal, ex: 13.5 = 13h30]", 12.0)
    tilt     = _ask("Inclinacao atual do painel [graus]", 23.5)
    azimute  = _ask("Azimute atual do painel [0=Sul, +L/-O]", 0.0)
    albedo   = _ask("Albedo da superficie [0.2=solo comum]", 0.2)

    print()
    print("  Fonte de irradiancia:")
    print("  [1] Open-Meteo  (recomendado, automatico)")
    print("  [2] NASA POWER  (fallback)")
    print("  [3] Inserir GHI manualmente")
    src = input("  Escolha [1/2/3]: ").strip()

    GHI = None
    modo_api = "openmeteo"
    if src == "2":
        modo_api = "nasapower"
    elif src == "3":
        GHI = _ask("GHI [W/m2]", 800.0)
        modo_api = "none"

    return {
        "lat": lat, "lon": lon, "tzone": tzone,
        "no_day": no_day, "hora_leg": hora_leg,
        "tilt": tilt, "azimute": azimute, "albedo": albedo,
        "GHI": GHI, "nome_local": nome_local,
        "_modo_api": modo_api,
    }


# ─────────────────────────────────────────────────────────────────────────────
#  PONTO DE ENTRADA
# ─────────────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Fator de Transposicao Solar - Hay + Erbs v2.0",
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("--xlsx",   type=str,   help="Caminho para planilha .xlsx")
    parser.add_argument("--cidade", type=str,   help="Nome da cidade (geocoding)")
    parser.add_argument("--lat",    type=float, help="Latitude [graus]")
    parser.add_argument("--lon",    type=float, help="Longitude [graus]")
    parser.add_argument("--tzone",  type=int,   help="Fuso horario GMT")
    parser.add_argument("--dia",    type=int,   help="Dia do ano [1-365]")
    parser.add_argument("--hora",   type=float, help="Hora legal [decimal]")
    parser.add_argument("--tilt",   type=float, help="Inclinacao do painel [graus]")
    parser.add_argument("--azim",   type=float, help="Azimute do painel [graus]")
    parser.add_argument("--ghi",    type=float, help="GHI manual [W/m2]")
    parser.add_argument("--albedo", type=float, default=0.2, help="Albedo [0-1]")
    parser.add_argument("--api",    type=str,   default="auto",
                        choices=["auto", "openmeteo", "nasapower", "none"],
                        help="Fonte de irradiancia (padrao=auto)")
    args = parser.parse_args()

    hoje     = datetime.date.today()
    base     = {}
    modo_api = args.api
    nome_saida = "FT_Solar"

    # ── Modo planilha ────────────────────────────────────────────────────────
    if args.xlsx:
        print(f"  Lendo: {args.xlsx}")
        try:
            base = ler_planilha(args.xlsx)
            nome_saida = os.path.splitext(os.path.basename(args.xlsx))[0]
        except Exception as e:
            print(f"  Erro ao ler planilha: {e}")
            sys.exit(1)

    # ── Modo cidade ──────────────────────────────────────────────────────────
    elif args.cidade:
        print(f"  Geocodificando '{args.cidade}'...")
        res = geocodificar(args.cidade)
        if not res:
            print("  Cidade nao encontrada. Use --lat/--lon.")
            sys.exit(1)
        lat, lon, display = res
        nome_local = display.split(",")[0]
        print(f"  {nome_local}  (lat={lat:.4f}, lon={lon:.4f})")
        tzone = args.tzone if args.tzone is not None else round(lon / 15)
        base = {
            "lat": lat, "lon": lon, "tzone": tzone,
            "no_day":   args.dia  or hoje.timetuple().tm_yday,
            "hora_leg": args.hora or 12.0,
            "tilt":     args.tilt or 23.5,
            "azimute":  args.azim or 0.0,
            "albedo":   args.albedo,
            "GHI":      args.ghi,
            "nome_local": nome_local,
        }
        nome_saida = args.cidade.replace(" ", "_")

    # ── Modo coordenadas diretas ─────────────────────────────────────────────
    elif args.lat is not None and args.lon is not None:
        tzone = args.tzone if args.tzone is not None else round(args.lon / 15)
        base = {
            "lat": args.lat, "lon": args.lon, "tzone": tzone,
            "no_day":   args.dia  or hoje.timetuple().tm_yday,
            "hora_leg": args.hora or 12.0,
            "tilt":     args.tilt or 23.5,
            "azimute":  args.azim or 0.0,
            "albedo":   args.albedo,
            "GHI":      args.ghi,
            "nome_local": f"Lat={args.lat:.2f} Lon={args.lon:.2f}",
        }
        nome_saida = f"FT_lat{args.lat}_lon{args.lon}"

    # ── Modo interativo ──────────────────────────────────────────────────────
    else:
        base     = input_interativo()
        modo_api = base.pop("_modo_api", "auto")
        nome_saida = str(base.get("nome_local", "FT_Solar")).replace(" ", "_")

    # Sobrescreve campos se passados via CLI (combina modos)
    if args.dia:    base["no_day"]   = args.dia
    if args.hora:   base["hora_leg"] = args.hora
    if args.tilt:   base["tilt"]     = args.tilt
    if args.azim:   base["azimute"]  = args.azim
    if args.ghi:    base["GHI"]      = args.ghi
    if args.albedo: base["albedo"]   = args.albedo

    # ── Completar, calcular e gerar ──────────────────────────────────────────
    params = montar_params(base, modo_api)

    print(f"  Calculando matriz FT (37x19 = 703 combinacoes)...")
    azimutes, inclinacoes, ft_matrix = calcular_matriz_ft(params)

    caminho_img = f"{nome_saida}_FT_map.png"
    print("  Gerando grafico...")
    ft_atual, best_val, best_tilt, best_az = gerar_grafico(
        params, azimutes, inclinacoes, ft_matrix, caminho_img
    )

    imprimir_relatorio(params, ft_atual, best_val, best_tilt, best_az, caminho_img)

    try:
        import subprocess, platform
        if platform.system() == "Windows":
            os.startfile(caminho_img)
        elif platform.system() == "Darwin":
            subprocess.run(["open", caminho_img])
        else:
            subprocess.run(["xdg-open", caminho_img])
    except Exception:
        pass

    return caminho_img


if __name__ == "__main__":
    main()
